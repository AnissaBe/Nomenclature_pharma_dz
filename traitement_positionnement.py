import pandas as pd
import re
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def traiter_positionnement(uploaded_file):
    """
    Prend un fichier CSV uploadé et retourne :
    - par_labo   : une ligne par laboratoire (format collègue)
    - par_code   : une ligne par code produit (notre format)
    - anomalies  : codes avec plusieurs DCI différentes
    """

    # ── Chargement ───────────────────────────────────────────────
    try:
        df = pd.read_csv(
            uploaded_file,
            sep=None,
            engine="python",
            encoding="utf-8"
        )
    except UnicodeDecodeError:
        uploaded_file.seek(0)
        df = pd.read_csv(
            uploaded_file,
            sep=None,
            engine="python",
            encoding="latin-1"
        )

    # ── Supprimer colonnes sans nom (Unnamed) ────────────────────
    df = df.loc[:, ~df.columns.str.contains("^Unnamed")]

    # ── Supprimer première colonne N° si elle existe ──────────────
    if df.columns[0] in ["N°", "N°ENREGISTREMENT", "N\u00b0", "N\u00b0ENREGISTREMENT"]:
        df = df.drop(columns=df.columns[0])

    df = df.reset_index(drop=True)

    # ── Noms des colonnes clés ───────────────────────────────────
    COL_CODE   = "CODE"
    COL_DCI    = "DENOMINATION COMMUNE INTERNATIONALE"
    COL_FORME  = "FORME"
    COL_DOSAGE = "DOSAGE"
    COL_COND   = "CONDITIONNEMENT"
    COL_LABO   = "LABORATOIRES DETENTEUR DE LA DECISION D'ENREGISTREMENT"
    COL_PAYS   = "PAYS DU LABORATOIRE DETENTEUR DE LA DECISION D'ENREGISTREMENT"
    COL_STATUT = "STATUT"

    # ── Nettoyage dosages ────────────────────────────────────────
    df[COL_DOSAGE] = df[COL_DOSAGE].apply(clean_dosage)

    # ════════════════════════════════════════════════════════════
    # ONGLET 1 — Par laboratoire (format collègue)
    # ════════════════════════════════════════════════════════════
    tableau = df[[
        COL_CODE, COL_DCI, COL_FORME, COL_DOSAGE,
        COL_COND, COL_LABO, COL_PAYS, COL_STATUT
    ]].copy()

    tableau["FABRIQUE"] = tableau[COL_STATUT].apply(lambda x: "R1" if str(x).strip().upper() == "F" else "")
    tableau["IMPORTE"]  = tableau[COL_STATUT].apply(lambda x: "R1" if str(x).strip().upper() == "I" else "")
    tableau = tableau.drop(columns=[COL_STATUT])

    tableau = tableau.rename(columns={
        COL_DCI:   "DCI",
        COL_COND:  "COND.",
        COL_LABO:  "LABORATOIRE",
        COL_PAYS:  "PAYS"
    })

    # Lignes TOTAL par CODE
    totaux = []
    for code, group in tableau.groupby(COL_CODE, sort=False):
        nb_f = (group["FABRIQUE"] == "R1").sum()
        nb_i = (group["IMPORTE"]  == "R1").sum()
        total = nb_f + nb_i
        totaux.append({
            COL_CODE:      f"Total {code}",
            "DCI":         "",
            COL_FORME:     "",
            COL_DOSAGE:    "",
            "COND.":       "",
            "LABORATOIRE": "",
            "PAYS":        "",
            "FABRIQUE":    f"R{nb_f}"  if nb_f > 0 else "",
            "IMPORTE":     f"R{nb_i}"  if nb_i > 0 else "",
            "TOTAL":       f"R{total}"
        })

    tableau["TOTAL"] = ""
    totaux_df = pd.DataFrame(totaux)

    result = []
    for code, group in tableau.groupby(COL_CODE, sort=False):
        result.append(group)
        total_row = totaux_df[totaux_df[COL_CODE] == f"Total {code}"]
        result.append(total_row)

    # Ligne TOTAL général
    nb_f_total = (tableau["FABRIQUE"] == "R1").sum()
    nb_i_total = (tableau["IMPORTE"]  == "R1").sum()
    total_general = pd.DataFrame([{
        COL_CODE:      "TOTAL",
        "DCI":         "",
        COL_FORME:     "",
        COL_DOSAGE:    "",
        "COND.":       "",
        "LABORATOIRE": "",
        "PAYS":        "",
        "FABRIQUE":    f"R{nb_f_total}",
        "IMPORTE":     f"R{nb_i_total}",
        "TOTAL":       f"R{nb_f_total + nb_i_total}"
    }])
    result.append(total_general)

    par_labo = pd.concat(result, ignore_index=True)

    # ════════════════════════════════════════════════════════════
    # ONGLET 2 — Par code (notre format)
    # ════════════════════════════════════════════════════════════
    par_code = df.groupby(COL_CODE).agg(
        DCI=(COL_DCI, "first"),
        FORME=(COL_FORME, "first"),
        DOSAGE=(COL_DOSAGE, "first"),
        **{"COND.": (COL_COND, "first")},
        LABORATOIRES=(COL_LABO, lambda x: ", ".join(x.dropna().unique())),
        PAYS=(COL_PAYS, lambda x: ", ".join(x.dropna().unique())),
        FABRIQUE=(COL_STATUT, lambda x: f"R{(x.str.strip().str.upper()=='F').sum()}"
                  if (x.str.strip().str.upper()=='F').sum() > 0 else ""),
        IMPORTE=(COL_STATUT, lambda x: f"R{(x.str.strip().str.upper()=='I').sum()}"
                 if (x.str.strip().str.upper()=='I').sum() > 0 else ""),
        TOTAL=(COL_STATUT, lambda x: f"R{len(x)}")
    ).reset_index()

    par_code = par_code.sort_values(COL_CODE).reset_index(drop=True)

    # ════════════════════════════════════════════════════════════
    # ONGLET 3 — Anomalies codes
    # ════════════════════════════════════════════════════════════
    anomalies_check = df.groupby(COL_CODE)[COL_DCI].nunique()
    anomalies_check = anomalies_check[anomalies_check > 1]

    rows = []
    for code in anomalies_check.index:
        dcis = df[df[COL_CODE] == code][COL_DCI].unique()
        row = {COL_CODE: code}
        for i, dci in enumerate(dcis):
            row[f"DCI_{i+1}"] = dci
        rows.append(row)

    anomalies = pd.DataFrame(rows).fillna("") if rows else pd.DataFrame()

    return par_labo, par_code, anomalies


# ═══════════════════════════════════════════════════════════════════
# FORMATAGE EXCEL
# ═══════════════════════════════════════════════════════════════════

def formater_feuille(ws, df, col_dci="DCI"):
    """Applique le formatage couleur et largeur à une feuille Excel."""

    # ── Couleurs ─────────────────────────────────────────────────
    BLEU_FONCE  = PatternFill("solid", fgColor="1A1A2E")  # en-têtes
    BLEU_CIEL   = PatternFill("solid", fgColor="BDD7EE")  # colonne CODE
    GRIS_CLAIR  = PatternFill("solid", fgColor="D9D9D9")  # lignes Total [CODE]
    GRIS_FONCE  = PatternFill("solid", fgColor="A6A6A6")  # ligne TOTAL général

    BLANC       = Font(color="FFFFFF", bold=True)
    NOIR        = Font(color="000000")
    NOIR_GRAS   = Font(color="000000", bold=True)

    # ── En-têtes (ligne 1) ───────────────────────────────────────
    for cell in ws[1]:
        cell.fill = BLEU_FONCE
        cell.font = BLANC
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Trouver index colonne CODE ────────────────────────────────
    col_code_idx = None
    col_dci_idx  = None
    for i, col in enumerate(df.columns, 1):
        if col == "CODE":
            col_code_idx = i
        if col == col_dci:
            col_dci_idx = i

    # ── Lignes de données ─────────────────────────────────────────
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        code_val = str(ws.cell(row=row_idx, column=1).value or "")

        # Déterminer le style de la ligne
        if code_val == "TOTAL":
            fill = GRIS_FONCE
            font = NOIR_GRAS
        elif code_val.startswith("Total "):
            fill = GRIS_CLAIR
            font = NOIR_GRAS
        else:
            fill = None
            font = NOIR

        for cell in row:
            if fill:
                cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(vertical="center", wrap_text=False)

        # Colonne CODE → bleu ciel (sauf lignes Total et TOTAL)
        if col_code_idx and not code_val.startswith("Total") and code_val != "TOTAL":
            code_cell = ws.cell(row=row_idx, column=col_code_idx)
            code_cell.fill = BLEU_CIEL
            code_cell.font = NOIR

    # ── Largeur des colonnes ──────────────────────────────────────
    for i, col in enumerate(df.columns, 1):
        col_letter = get_column_letter(i)
        if col == col_dci or col == "LABORATOIRES" or col == "LABORATOIRE":
            ws.column_dimensions[col_letter].width = 40
        elif col == "CODE":
            ws.column_dimensions[col_letter].width = 12
        elif col in ["FABRIQUE", "IMPORTE", "TOTAL"]:
            ws.column_dimensions[col_letter].width = 10
        elif col in ["PAYS"]:
            ws.column_dimensions[col_letter].width = 15
        elif col in ["FORME"]:
            ws.column_dimensions[col_letter].width = 25
        elif col in ["DOSAGE"]:
            ws.column_dimensions[col_letter].width = 18
        elif col in ["COND."]:
            ws.column_dimensions[col_letter].width = 20
        else:
            # Auto-ajustement pour les autres colonnes
            max_len = max(
                (len(str(ws.cell(row=r, column=i).value or "")) for r in range(1, ws.max_row + 1)),
                default=10
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

    # ── Figer la première ligne ───────────────────────────────────
    ws.freeze_panes = "A2"


# ── Nettoyage dosages ─────────────────────────────────────────────
def clean_dosage(s):
    if pd.isna(s):
        return s
    s = str(s).strip().upper()
    unites = r'(MG/ML|G/ML|MG/G|MG/L|UI/ML|UI/MG|MCG/ML|MG/CM2|MMOL/ML|MEQ/ML|MG|MCG|UI|GR|ML|G|%)'
    s = re.sub(r'(\d[\d,\.]*)\s+' + unites, r'\1\2', s)
    s = re.sub(r'\(\s+', '(', s)
    s = re.sub(r'\s+\)', ')', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s
    """
    Prend un fichier CSV uploadé et retourne :
    - par_labo   : une ligne par laboratoire (format collègue)
    - par_code   : une ligne par code produit (notre format)
    - anomalies  : codes avec plusieurs DCI différentes
    """

    # ── Chargement ───────────────────────────────────────────────
    try:
        df = pd.read_csv(
            uploaded_file,
            sep=None,
            engine="python",
            encoding="utf-8"
        )
    except UnicodeDecodeError:
        uploaded_file.seek(0)
        df = pd.read_csv(
            uploaded_file,
            sep=None,
            engine="python",
            encoding="latin-1"
        )

    # ── Supprimer colonnes sans nom (Unnamed) ────────────────────
    df = df.loc[:, ~df.columns.str.contains("^Unnamed")]

    # ── Supprimer première colonne N° si elle existe ──────────────
    if df.columns[0] in ["N°", "N°ENREGISTREMENT", "N\u00b0", "N\u00b0ENREGISTREMENT"]:
        df = df.drop(columns=df.columns[0])

    df = df.reset_index(drop=True)

    # ── Noms des colonnes clés ───────────────────────────────────
    COL_CODE   = "CODE"
    COL_DCI    = "DENOMINATION COMMUNE INTERNATIONALE"
    COL_FORME  = "FORME"
    COL_DOSAGE = "DOSAGE"
    COL_COND   = "CONDITIONNEMENT"
    COL_LABO   = "LABORATOIRES DETENTEUR DE LA DECISION D'ENREGISTREMENT"
    COL_PAYS   = "PAYS DU LABORATOIRE DETENTEUR DE LA DECISION D'ENREGISTREMENT"
    COL_STATUT = "STATUT"

    # ── Nettoyage dosages ────────────────────────────────────────
    df[COL_DOSAGE] = df[COL_DOSAGE].apply(clean_dosage)

    # ════════════════════════════════════════════════════════════
    # ONGLET 1 — Par laboratoire (format collègue)
    # ════════════════════════════════════════════════════════════
    tableau = df[[
        COL_CODE, COL_DCI, COL_FORME, COL_DOSAGE,
        COL_COND, COL_LABO, COL_PAYS, COL_STATUT
    ]].copy()

    tableau["FABRIQUE"] = tableau[COL_STATUT].apply(lambda x: "R1" if str(x).strip().upper() == "F" else "")
    tableau["IMPORTE"]  = tableau[COL_STATUT].apply(lambda x: "R1" if str(x).strip().upper() == "I" else "")
    tableau = tableau.drop(columns=[COL_STATUT])

    tableau = tableau.rename(columns={
        COL_DCI:   "DCI",
        COL_COND:  "COND.",
        COL_LABO:  "LABORATOIRE",
        COL_PAYS:  "PAYS"
    })

    # Lignes TOTAL par CODE
    totaux = []
    for code, group in tableau.groupby(COL_CODE, sort=False):
        nb_f = (group["FABRIQUE"] == "R1").sum()
        nb_i = (group["IMPORTE"]  == "R1").sum()
        total = nb_f + nb_i
        totaux.append({
            COL_CODE:      f"Total {code}",
            "DCI":         "",
            COL_FORME:     "",
            COL_DOSAGE:    "",
            "COND.":       "",
            "LABORATOIRE": "",
            "PAYS":        "",
            "FABRIQUE":    f"R{nb_f}"  if nb_f > 0 else "",
            "IMPORTE":     f"R{nb_i}"  if nb_i > 0 else "",
            "TOTAL":       f"R{total}"
        })

    tableau["TOTAL"] = ""
    totaux_df = pd.DataFrame(totaux)

    result = []
    for code, group in tableau.groupby(COL_CODE, sort=False):
        result.append(group)
        total_row = totaux_df[totaux_df[COL_CODE] == f"Total {code}"]
        result.append(total_row)

    # Ligne TOTAL général
    nb_f_total = (tableau["FABRIQUE"] == "R1").sum()
    nb_i_total = (tableau["IMPORTE"]  == "R1").sum()
    total_general = pd.DataFrame([{
        COL_CODE:      "TOTAL",
        "DCI":         "",
        COL_FORME:     "",
        COL_DOSAGE:    "",
        "COND.":       "",
        "LABORATOIRE": "",
        "PAYS":        "",
        "FABRIQUE":    f"R{nb_f_total}",
        "IMPORTE":     f"R{nb_i_total}",
        "TOTAL":       f"R{nb_f_total + nb_i_total}"
    }])
    result.append(total_general)

    par_labo = pd.concat(result, ignore_index=True)

    # ════════════════════════════════════════════════════════════
    # ONGLET 2 — Par code (notre format)
    # ════════════════════════════════════════════════════════════
    par_code = df.groupby(COL_CODE).agg(
        DCI=(COL_DCI, "first"),
        FORME=(COL_FORME, "first"),
        DOSAGE=(COL_DOSAGE, "first"),
        **{"COND.": (COL_COND, "first")},
        LABORATOIRES=(COL_LABO, lambda x: ", ".join(x.dropna().unique())),
        PAYS=(COL_PAYS, lambda x: ", ".join(x.dropna().unique())),
        FABRIQUE=(COL_STATUT, lambda x: f"R{(x.str.strip().str.upper()=='F').sum()}"
                  if (x.str.strip().str.upper()=='F').sum() > 0 else ""),
        IMPORTE=(COL_STATUT, lambda x: f"R{(x.str.strip().str.upper()=='I').sum()}"
                 if (x.str.strip().str.upper()=='I').sum() > 0 else ""),
        TOTAL=(COL_STATUT, lambda x: f"R{len(x)}")
    ).reset_index()

    par_code = par_code.sort_values(COL_CODE).reset_index(drop=True)

    # ════════════════════════════════════════════════════════════
    # ONGLET 3 — Anomalies codes
    # ════════════════════════════════════════════════════════════
    anomalies_check = df.groupby(COL_CODE)[COL_DCI].nunique()
    anomalies_check = anomalies_check[anomalies_check > 1]

    rows = []
    for code in anomalies_check.index:
        dcis = df[df[COL_CODE] == code][COL_DCI].unique()
        row = {COL_CODE: code}
        for i, dci in enumerate(dcis):
            row[f"DCI_{i+1}"] = dci
        rows.append(row)

    anomalies = pd.DataFrame(rows).fillna("") if rows else pd.DataFrame()

    return par_labo, par_code, anomalies


# ── Nettoyage dosages ─────────────────────────────────────────────
def clean_dosage(s):
    if pd.isna(s):
        return s
    s = str(s).strip().upper()
    unites = r'(MG/ML|G/ML|MG/G|MG/L|UI/ML|UI/MG|MCG/ML|MG/CM2|MMOL/ML|MEQ/ML|MG|MCG|UI|GR|ML|G|%)'
    s = re.sub(r'(\d[\d,\.]*)\s+' + unites, r'\1\2', s)
    s = re.sub(r'\(\s+', '(', s)
    s = re.sub(r'\s+\)', ')', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s